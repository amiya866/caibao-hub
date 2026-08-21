# -*- coding: utf-8 -*-
"""把 _pb_fill.json 的铅 mine/refined 数据回填进 excel/铅.xlsx（补缺不覆盖）。"""
import json, sys
import openpyxl

sys.stdout.reconfigure(encoding="utf-8")

XLSX = r"D:\拷贝文件\E\永安\铅\全球铅企季度产量梳理.xlsx"
FILL = r"D:\大胖鱼\财报汇总\_pb_fill.json"

# 标准 26 列：期间 -> 列号
COLMAP = {}
for i, (y, cols) in enumerate({
    "2023": (4, 5, 6, 7, 8), "2024": (9, 10, 11, 12, 13), "2025": (14, 15, 16, 17, 18),
}.items()):
    for q in range(1, 5):
        COLMAP[f"{y}Q{q}"] = cols[q - 1]
    COLMAP[y] = cols[4]
COLMAP["2026Q1"] = 20
COLMAP["2026Q2"] = 22
NOTE_COL = 26  # Z=备注

fill = json.load(open(FILL, encoding="utf-8"))
wb = openpyxl.load_workbook(XLSX)  # 保留公式（同比列/总计行）

stats = {"cells": 0, "skipped_existing": 0, "rows": set(), "missing_rows": []}
for sec_key, sheet_name in [("mine", "铅精矿·季度产量"), ("refined", "精炼铅·季度产量")]:
    ws = wb[sheet_name]
    # 建立 (公司||项目) -> 行号 索引
    rowidx = {}
    for r in range(2, ws.max_row + 1):
        a = ws.cell(row=r, column=1).value
        c = ws.cell(row=r, column=3).value
        if a and a != "总计":
            rowidx[f"{str(a).strip()}||{str(c).strip()}"] = r
    for key, rec in fill[sec_key].items():
        r = rowidx.get(key)
        if r is None:
            stats["missing_rows"].append((sheet_name, key))
            continue
        stats["rows"].add((sheet_name, key))
        for period, col in COLMAP.items():
            if period not in rec:
                continue
            cell = ws.cell(row=r, column=col)
            cur = cell.value
            if isinstance(cur, (int, float)):  # 已有数值，不覆盖
                stats["skipped_existing"] += 1
                continue
            cell.value = rec[period]
            stats["cells"] += 1
        # 备注列：保留原有口径备注，追加来源
        note_cell = ws.cell(row=r, column=NOTE_COL)
        src = rec.get("source", "").strip()
        old = (note_cell.value or "").strip()
        if src and src not in old:
            note_cell.value = (old + "；来源：" + src) if old else ("来源：" + src)

# 更新日志追加一行
log = wb["更新日志"]
lr = log.max_row + 1
log.cell(row=lr, column=1, value="2026-08-21")
log.cell(row=lr, column=2, value="回填产量数据：铅精矿 12 行（Boliden两矿/MMG两矿/South32 Cannington/Vedanta年度/Peñoles/Volcan/Fresnillo/BHP Antamina/Newmont/SCCO年度）+ 精炼铅 2 行（Peñoles Torreon、HZL），2023Q1-2026Q2")
log.cell(row=lr, column=3, value="各公司官方季报/年报（2026-08-18 核查底稿 _pb_fill.json）")

wb.save(XLSX)
print("写入单元格:", stats["cells"], "| 跳过已有值:", stats["skipped_existing"])
print("命中行:", len(stats["rows"]))
for s, k in sorted(stats["rows"]):
    print("  ", s, k)
if stats["missing_rows"]:
    print("!! 未匹配行:", stats["missing_rows"])
