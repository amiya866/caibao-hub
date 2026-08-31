# -*- coding: utf-8 -*-
"""财报平台欠账守望者 v2（2026-08-23 建，2026-08-31 v2 升级）
每周一 09:11（Windows 任务 CaibaoWatch）：
  ① 扫 8 品种真库：上季有数但当季空 = 疑似「已披露未入库」；
  ② 披露日历交叉核对：build_site.py 日历中日期已过期的公司，真库当季仍空 = 「已披露待核·未入库」（带逾期天数）；
     ——网站披露日历的状态已与真库联动（已入库/已披露待核·未入库），本清单是它的每周自检；
  ③ 有欠账就弹窗提醒（叫 Kimi 来补）；提醒信息速递每周扫描。
CUR_Q 自动推算（_curq.py：Q1 季 4/20、Q2 季 8/10、Q3 季 10/25、Q4 季次年 2/20 起核），无需手工调。
手动跑: python _caibao_watch.py
"""
import ast
import sys
from datetime import date
from pathlib import Path

import openpyxl

sys.path.insert(0, str(Path(__file__).resolve().parent))
from _curq import cur_check_quarter, company_match  # noqa: E402

sys.stdout.reconfigure(encoding="utf-8", errors="replace")

BASE = Path(r"D:\Kimi\caibao-hub")
FILES = {
    "铜": r"D:\Kimi\金属总网\网站构建\财报汇总\铜\全球铜企季度产量梳理.xlsx",
    "铝": r"D:\Kimi\金属总网\网站构建\财报汇总\铝\全球铝企季度产量梳理.xlsx",
    "铅": r"D:\Kimi\金属总网\网站构建\财报汇总\铅\全球铅企季度产量梳理.xlsx",
    "锌": r"D:\Kimi\金属总网\网站构建\财报汇总\锌\全球锌企季度产量梳理.xlsx",
    "镍": r"D:\Kimi\金属总网\网站构建\财报汇总\镍\全球镍企季度产量梳理.xlsx",
    "锂": r"D:\Kimi\金属总网\网站构建\财报汇总\锂\全球锂企季度产量梳理.xlsx",
    "硅": r"D:\Kimi\金属总网\网站构建\财报汇总\硅\全球硅企季度产量梳理.xlsx",
    "锡": r"D:\Kimi\金属总网\网站构建\财报汇总\锡\海外主要公司产量.xlsx",
}
OUT = BASE / "_待核清单.md"
CUR_Q, PREV_Q = cur_check_quarter()

CAL_VAR2COMM = {"TIN": "锡", "ZINC": "锌", "ALUMINUM": "铝", "NICKEL": "镍",
                "COPPER": "铜", "LITHIUM": "锂", "SILICON": "硅", "LEAD": "铅"}


def load_calendars():
    """ast 解析 build_site.py 的披露日历常量（不执行整个构建脚本）。"""
    tree = ast.parse((BASE / "build_site.py").read_text(encoding="utf-8"))
    cals = {}
    for node in tree.body:
        if isinstance(node, ast.Assign) and isinstance(node.targets[0], ast.Name):
            var = node.targets[0].id
            if var.endswith("_CALENDAR") and var[:-9] in CAL_VAR2COMM:
                cals[CAL_VAR2COMM[var[:-9]]] = ast.literal_eval(node.value)
    return cals


def sheet_view(path):
    """每 sheet → (CUR_Q 列号, PREV_Q 列号, {行名: 行 tuple})；无 CUR_Q 表头的 sheet 跳过。"""
    views = []
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    for ws in wb:
        if "日志" in ws.title:
            continue
        hdr = None
        hdr_row = 0
        for i, r in enumerate(ws.iter_rows(min_row=1, max_row=5, values_only=True), start=1):
            if r and any(str(v).strip() == CUR_Q for v in r if v):
                hdr = r
                hdr_row = i
                break
        if not hdr:
            continue
        cols = {str(v).strip(): j for j, v in enumerate(hdr) if v}
        if CUR_Q not in cols or PREV_Q not in cols:
            continue
        rows = {}
        for r in ws.iter_rows(min_row=hdr_row + 1, values_only=True):
            name = r[0]
            if not name or "总计" in str(name) or "合计" in str(name):
                continue
            rows.setdefault(str(name), []).append(r)
        views.append((ws.title, cols, rows))
    wb.close()
    return views


def scan(path):
    """① 上季有数、当季空的公司行。"""
    owes = []
    try:
        views = sheet_view(path)
    except Exception as e:
        return [("(文件读取失败)", str(e)[:60], "", "")]
    for title, cols, rows in views:
        for name, rlist in rows.items():
            for r in rlist:
                cur, prev = r[cols[CUR_Q]], r[cols[PREV_Q]]
                if prev not in (None, "", "-") and cur in (None, ""):
                    owes.append((title, name, str(r[1] or ""), str(r[2] or "") if len(r) > 2 else ""))
    return owes


def cal_cross(comm, path, cal_entries):
    """② 日历已到期 × 真库当季空缺。返回 (已披露待核·未入库 list, 无对应行 list)。"""
    try:
        views = sheet_view(path)
    except Exception:
        return [], []
    row_names = [n for _, _, rows in views for n in rows]
    missing, norow = [], []
    for e in cal_entries:
        d = date.fromisoformat(e["date"])
        if d > date.today():
            continue
        row = company_match(e["company"], row_names)
        if row is None:
            norow.append((e["company"], e["date"], e["event"]))
            continue
        filled = False
        for _, cols, rows in views:
            if row in rows:
                for r in rows[row]:
                    v = r[cols[CUR_Q]]
                    if v not in (None, "", "-"):
                        filled = True
                        break
            if filled:
                break
        if not filled:
            missing.append((e["company"], e["date"], e["event"], (date.today() - d).days))
    return missing, norow


def popup(msg):
    import subprocess
    try:
        esc = msg.replace("'", "''").replace("\n", "`n")
        subprocess.Popen(["powershell", "-NoProfile", "-WindowStyle", "Hidden", "-Command",
                          "Add-Type -AssemblyName System.Windows.Forms;"
                          f"[System.Windows.Forms.MessageBox]::Show('{esc}','财报平台欠账提醒','OK','Warning')"])
    except Exception:
        pass


def main():
    lines = [f"# 财报平台待核清单（{date.today()} 自动生成，守望者 _caibao_watch.py v2）",
             f"\n当前应核季度={CUR_Q}（自动推算）。①={PREV_Q} 有数但 {CUR_Q} 空；②=披露日历已到期但 {CUR_Q} 真库仍空（已披露待核·未入库）。\n"]
    total = 0
    for comm, path in FILES.items():
        owes = scan(path)
        total += len(owes)
        lines.append(f"\n## {comm}（{len(owes)} 行）")
        for sheet, name, country, proj in owes:
            lines.append(f"- [{sheet}] {name} {country} {proj}".rstrip())
    # ② 日历交叉核对
    cals = load_calendars()
    lines.append(f"\n---\n\n# ② 披露日历交叉核对（已披露待核·未入库）")
    cal_total = 0
    for comm, entries in cals.items():
        missing, norow = cal_cross(comm, FILES[comm], entries)
        cal_total += len(missing)
        if missing or norow:
            lines.append(f"\n## {comm}（未入库 {len(missing)} / 无对应行 {len(norow)}）")
            for name, d, event, days in missing:
                lines.append(f"- **{name}**｜{d}｜{event}｜逾期 {days} 天")
            for name, d, event in norow:
                lines.append(f"- {name}｜{d}｜{event}｜真库无对应行（需手工核对或加别名）")
    lines.append(f"\n② 合计未入库 {cal_total} 条。")
    OUT.write_text("\n".join(lines), encoding="utf-8")
    print(f"欠账① {total} 行 + 日历② {cal_total} 条 → {OUT}")
    if total or cal_total:
        popup(f"财报平台欠账：①空缺 {total} 行 + ②已披露未入库 {cal_total} 条（{CUR_Q}）。\n清单：{OUT}\n叫 Kimi：「按待核清单更新财报平台」。\n\n另：信息速递每周扫描（印尼镍/铝讯/阿拉丁/爱择/抖音传言）也别忘。")


if __name__ == "__main__":
    main()
